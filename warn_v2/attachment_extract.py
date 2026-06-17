"""Best-effort field extraction from non-PDF WARN notice attachments.

GA's TCSG entry pages attach the WARN notice as a ``gk-download`` file that is
often *not* a PDF — frequently a Word document or a spreadsheet, occasionally
HTML/CSV/plain text. PDFs go through :mod:`warn_v2.pdf_extract`; this module
turns the other formats into text and feeds it to the *same* ``_parse_text``
field parser, so the recipient-aware worksite selection in ``_choose_city_zip``
(a WARN letter is addressed to state officials, so naive first-match pins the
recipient) applies uniformly across formats.

All extraction is best-effort: any failure yields ``{}`` rather than raising.
"""
from __future__ import annotations

import csv
import io
import logging
import zipfile

log = logging.getLogger(__name__)

_ZIP_MAGIC = b"PK\x03\x04"
# OOXML WordprocessingML namespace — paragraph (<w:p>) and text-run (<w:t>) tags.
_W_NS = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"


def extract_attachment_fields(
    content: bytes,
    content_type: str | None,
    filename: str | None,
    state: str | None,
) -> dict:
    """Extract WARN fields from an attachment of any supported type.

    Routes by magic bytes / content-type / filename to a text extractor, then
    parses fields with the shared ``_parse_text``. PDFs delegate to
    ``extract_warn_fields`` (pdfplumber + OCR fallback). Returns ``{}`` on any
    failure or an unsupported type.
    """
    try:
        kind = _classify(content, content_type, filename)
        if kind == "pdf":
            from warn_v2.pdf_extract import extract_warn_fields

            return extract_warn_fields(content, state)
        text = _to_text(kind, content)
        if not text.strip():
            return {}
        from warn_v2.pdf_extract import _parse_text

        return _parse_text(text, state)
    except Exception as e:  # never raise — a bad attachment must not crash a run
        log.debug("attachment_extract: failed (content_type=%r): %s", content_type, e)
        return {}


def _classify(content: bytes, content_type: str | None, filename: str | None) -> str | None:
    """Best-effort attachment type: pdf|docx|xlsx|csv|html|text|None.

    Trusts magic bytes first (content-type/filename from TCSG are unreliable),
    then content-type, then filename extension.
    """
    ct = (content_type or "").lower()
    name = (filename or "").lower()

    if content[:4] == b"%PDF" or "pdf" in ct or name.endswith(".pdf"):
        return "pdf"
    if content[:4] == _ZIP_MAGIC:
        ooxml = _ooxml_kind(content)
        if ooxml:
            return ooxml
    if "wordprocessingml" in ct or name.endswith(".docx"):
        return "docx"
    if "spreadsheetml" in ct or name.endswith(".xlsx"):
        return "xlsx"
    if "csv" in ct or name.endswith(".csv"):
        return "csv"
    if "html" in ct or name.endswith((".html", ".htm")):
        return "html"
    if ct.startswith("text/") or name.endswith(".txt"):
        return "text"
    return None


def _ooxml_kind(content: bytes) -> str | None:
    """A zip-magic file is OOXML docx/xlsx — disambiguate by its member names."""
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as zf:
            names = set(zf.namelist())
    except Exception:
        return None
    if "word/document.xml" in names:
        return "docx"
    if "xl/workbook.xml" in names:
        return "xlsx"
    return None


def _to_text(kind: str | None, content: bytes) -> str:
    if kind == "docx":
        return _docx_to_text(content)
    if kind == "xlsx":
        return _xlsx_to_text(content)
    if kind == "csv":
        return _csv_to_text(content)
    if kind == "html":
        return _html_to_text(content)
    if kind == "text":
        return content.decode("utf-8", errors="replace")
    return ""


def _docx_to_text(content: bytes) -> str:
    """Join Word paragraph text from ``word/document.xml`` (stdlib zip + lxml).

    One line per ``<w:p>`` so ``pdf_extract._choose_city_zip``'s line windows
    still see a city on its own line under a street/recipient line.
    """
    from lxml import etree

    with zipfile.ZipFile(io.BytesIO(content)) as zf:
        xml = zf.read("word/document.xml")
    root = etree.fromstring(xml)
    lines: list[str] = []
    for para in root.iter(f"{_W_NS}p"):
        runs = [t.text for t in para.iter(f"{_W_NS}t") if t.text]
        if runs:
            lines.append("".join(runs))
    return "\n".join(lines)


def _xlsx_to_text(content: bytes) -> str:
    """Join spreadsheet cells row-by-row (first sheet) — best-effort for rosters."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        ws = wb.active
        if ws is None:
            return ""
        lines: list[str] = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            if cells:
                lines.append(" ".join(cells))
        return "\n".join(lines)
    finally:
        wb.close()


def _csv_to_text(content: bytes) -> str:
    text = content.decode("utf-8", errors="replace")
    rows = csv.reader(io.StringIO(text))
    return "\n".join(" ".join(c for c in row if c) for row in rows)


def _html_to_text(content: bytes) -> str:
    from bs4 import BeautifulSoup

    return BeautifulSoup(content, "html.parser").get_text("\n")
